import sys
import time
import re
import os
import shutil
import json
from urllib.parse import urlparse, urljoin


import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# --- HTTP Session ---
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,tr;q=0.8",
})
REQUEST_TIMEOUT = 15

# Known generic email prefixes to prefer
PREFERRED_PREFIXES = ["info", "contact", "sales", "office", "marketing", "export", "support", "mail", "hello"]

# Pages to check for contact info
CONTACT_PATHS = [
    "/contact", "/contact-us", "/contactus",
    "/iletisim", "/bize-ulasin",
    "/about", "/about-us",
    "/company",
    "/support",
    "/tr/iletisim",
    "/en/contact",
]

# Stats
STATS = {
    "companies_processed": 0,
    "emails_added": 0,
    "phones_added": 0,
    "websites_added": 0,
    "remaining_no_email": 0,
    "remaining_no_phone": 0,
}


def load_checkpoint(checkpoint_file):
    if os.path.exists(checkpoint_file):
        with open(checkpoint_file, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_checkpoint(done_slugs, checkpoint_file):
    with open(checkpoint_file, "w", encoding="utf-8") as f:
        json.dump(list(done_slugs), f)


def clean_url(url):
    url = url.strip().strip('"').strip("'")
    if not url or url == "-":
        return ""
    if "@" in url and "." in url.split("@")[1] if "@" in url else False:
        return ""

    if url.startswith("http://") or url.startswith("https://"):
        parsed = urlparse(url)
        netloc = parsed.netloc
        if not netloc or " " in netloc or not netloc.count("."):
            return ""
        return url

    if url.startswith("www."):
        test = "https://" + url
        parsed = urlparse(test)
        if parsed.netloc and "." in parsed.netloc and " " not in parsed.netloc:
            return test
        return ""

    if "." in url and " " not in url:
        return "https://" + url

    return ""


def normalize_domain(url):
    try:
        parsed = urlparse(url)
        domain = parsed.netloc or parsed.path
        if domain.startswith("www."):
            domain = domain[4:]
        return domain.lower()
    except Exception:
        return ""


def get_base_url(url):
    try:
        parsed = urlparse(url)
        return f"{parsed.scheme}://{parsed.netloc}"
    except Exception:
        return url


def fetch_page(url, timeout=REQUEST_TIMEOUT):
    """Fetch a page with requests. Returns (status, html) or raises."""
    r = SESSION.get(url, timeout=timeout, allow_redirects=True)
    r.raise_for_status()
    content_type = r.headers.get("Content-Type", "")
    if "text/html" not in content_type and "application/xhtml" not in content_type:
        return r.status_code, ""
    return r.status_code, r.text


def extract_emails(html, domain=""):
    """Extract emails from HTML text."""
    # Find all emails
    emails = set()
    patterns = [
        r'[\w.+-]+@[\w-]+\.[\w.-]+',
    ]
    for p in patterns:
        found = re.findall(p, html)
        for e in found:
            e = e.strip().lower()
            # Filter invalid
            if not re.match(r'^[^@]+@[^@]+\.[^@]{2,}$', e):
                continue
            # Skip image extensions
            if e.endswith(('.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.css', '.js')):
                continue
            # Skip obvious non-emails
            if e.startswith(('@', '.')) or '..' in e:
                continue
            emails.add(e)
    return list(emails)


def extract_phones(text):
    """Extract phone numbers from text."""
    phones = set()
    patterns = [
        r'(?:\+?\d{1,3}[-.\s]?)?\(?\d{2,4}\)?[-.\s]?\d{2,4}[-.\s]?\d{2,4}[-.\s]?\d{2,4}',
        r'\+\d{1,3}[-.\s]?\d{1,4}[-.\s]?\d{2,4}[-.\s]?\d{2,4}[-.\s]?\d{2,4}',
    ]
    for p in patterns:
        found = re.findall(p, text)
        for ph in found:
            ph = ph.strip()
            digits = re.sub(r'\D', '', ph)
            if 8 <= len(digits) <= 15:
                phones.add(ph)
    return list(phones)


def search_company_website(company_name):
    """Search for company website using DuckDuckGo."""
    queries = [
        f'"{company_name}" official website',
        f'"{company_name}"',
        company_name,
    ]
    for query in queries:
        try:
            r = SESSION.post(
                "https://lite.duckduckgo.com/lite/",
                data={"q": query},
                timeout=10,
            )
            if r.status_code != 200:
                continue

            soup = BeautifulSoup(r.text, "lxml")
            for a in soup.select("a.result-link"):
                href = a.get("href", "")
                if href.startswith("http"):
                    return href
        except Exception:
            continue
    return ""


def find_contact_page_urls(base_url, soup):
    """Find likely contact page URLs from the homepage."""
    contact_urls = set()
    base = base_url.rstrip("/")

    # Add standard paths
    for path in CONTACT_PATHS:
        contact_urls.add(base + path)

    # Find links on page
    if soup:
        for a in soup.find_all("a", href=True):
            href = a["href"]
            text = a.get_text(strip=True).lower()
            link_lower = href.lower()
            # Check if link text or URL contains contact-related words
            keywords = ["contact", "iletisim", "bize", "ulasin", "about", "hakkimizda", "reach", "get in touch"]
            if any(k in text for k in keywords) or any(k in link_lower for k in keywords):
                full_url = urljoin(base, href)
                if full_url.startswith("http"):
                    contact_urls.add(full_url)

    return list(contact_urls)


def extract_from_soup(soup, base_url):
    """Extract emails and phones from a parsed page."""
    text = soup.get_text(separator=" ", strip=True) if soup else ""
    html = str(soup) if soup else ""

    emails = extract_emails(html)
    domain = normalize_domain(base_url)
    filtered = []
    for e in emails:
        # Skip generic/system emails
        if any(x in e for x in ["example.com", "domain.com", "yourdomain", "@email.com",
                                 "noreply", "no-reply", "donotreply", "wordpress", "localhost"]):
            continue
        if e.count('@') != 1:
            continue
        # Skip version number patterns like intl-tel-input@18.1.5, parent@5.4.6
        local, domain_part = e.split("@")
        domain_parts = domain_part.split(".")
        # If domain part has 2+ segments and all are numbers, it's a version
        if len(domain_parts) >= 2:
            all_digit_segments = all(p.isdigit() for p in domain_parts)
            if all_digit_segments:
                continue
        # Skip single characters before @
        if len(local) <= 1:
            continue
        # Prefer emails matching the website domain
        if domain and domain in domain_part:
            filtered.insert(0, e)
        else:
            filtered.append(e)
    emails = filtered

    phones = extract_phones(text)

    return emails, phones


def score_email(email):
    """Score email by prefix preference. Lower is better."""
    prefix = email.split("@")[0].lower()
    for i, p in enumerate(PREFERRED_PREFIXES):
        if prefix == p or prefix.startswith(p + "."):
            return i
    return len(PREFERRED_PREFIXES)


def pick_best_email(emails):
    """Pick the best email from a list."""
    if not emails:
        return ""
    # Remove duplicates
    unique = list(set(emails))
    if len(unique) == 1:
        return unique[0]
    # Sort by preference
    unique.sort(key=score_email)
    return unique[0]


def is_valid_email(e):
    return bool(re.match(r'^[^@]+@[^@]+\.[^@]{2,}$', e))


def enrich_company(name, existing_website, existing_phone):
    """Main enrichment function for a single company."""
    email = ""
    phone = existing_phone
    website = existing_website
    found_website = False

    # Önce anlamlı kelimeleri çıkar (tüm yöntemlerde doğrulama için)
    name_normalized = name.replace('İ', 'i').replace('I', 'i')
    name_normalized = name_normalized.replace('ı', 'i').replace('ş', 's').replace('Ş', 's')
    name_normalized = name_normalized.replace('ç', 'c').replace('Ç', 'c')
    name_normalized = name_normalized.replace('ğ', 'g').replace('Ğ', 'g')
    name_normalized = name_normalized.replace('ü', 'u').replace('Ü', 'u')
    name_normalized = name_normalized.replace('ö', 'o').replace('Ö', 'o')
    name_normalized = name_normalized.replace('â', 'a').replace('Â', 'a')
    name_normalized = name_normalized.replace('î', 'i').replace('Î', 'i')
    name_normalized = name_normalized.replace('û', 'u').replace('Û', 'u')
    name_normalized = name_normalized.lower()

    name_clean = re.sub(r'\s+(ANONIM|A\.S\.|A\.S|SAN\.|TIC\.|LTD\.|STI\.|LTD|STI|GMBH|SPA|AG|CO|KG|AS|NV|SCA|SE|GMBH|GMBH\s+CO\s+KG)\s*$', '', name_normalized, flags=re.IGNORECASE)
    name_clean = re.sub(r'[^\w\s]', ' ', name_clean)
    name_clean = name_clean.strip()

    raw_words = re.findall(r'[a-zA-Z0-9]+', name_clean)

    SHORT_WORDS = {"a", "an", "as", "at", "by", "co", "de", "do", "el", "en", "es", "et", "go",
                   "hi", "ic", "id", "if", "in", "is", "it", "la", "le", "lo", "me", "my",
                   "no", "of", "on", "or", "re", "se", "si", "so", "to", "un", "up", "us", "we",
                   "ve", "veya", "ile", "bir", "san", "tic", "ltd", "sti", "ltdti", "as",
                   "a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m",
                   "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z"}

    words = [w for w in raw_words if w not in SHORT_WORDS and len(w) > 1]

    # Step 1: If no website, search for it via DuckDuckGo
    if not website:
        print(f"    Searching for website...", end="")
        found_url = search_company_website(name)
        if found_url:
            try:
                r = SESSION.get(found_url, timeout=5)
                if r.status_code < 400 and "text/html" in r.headers.get("Content-Type", ""):
                    text_lower = r.text.lower()
                    meaningful_in_page = sum(1 for w in words[:5] if w in text_lower)
                    if words and meaningful_in_page >= min(2, len(words)):
                        website = found_url
                        found_website = True
                        print(f" found (verified): {website[:80]}")
                    else:
                        print(f" rejected: {found_url} (no company name match)")
                else:
                    print(f" rejected: {found_url} (no HTML)")
            except Exception as e:
                print(f" error: {e}")
        else:
            print(f" not found")

    # Step 2: If still no website, try domain guessing
    if not website:
        candidates = set()
        if words:
            candidates.add(words[0])
            candidates.add(''.join(words))
            if len(words) > 1:
                candidates.add(words[0] + words[1])
                candidates.add(words[0] + '-' + words[1])
            for w in words[1:]:
                candidates.add(w)

        for domain_base in candidates:
            for ext in [".com.tr", ".com", ".de", ".it", ".no", ".eu", ".net", ".org"]:
                test_url = f"https://{domain_base}{ext}"
                try:
                    r = SESSION.get(test_url, timeout=5)
                    if r.status_code < 400:
                        if "text/html" in r.headers.get("Content-Type", ""):
                            text_lower = r.text.lower()
                            meaningful_in_page = sum(1 for w in words[:5] if w in text_lower)
                            if words and meaningful_in_page < min(2, len(words)):
                                continue
                        print(f" found via domain guess: {test_url}")
                        website = test_url
                        found_website = True
                        break
                except Exception:
                    continue
            if website:
                break

    # Step 2: Visit website and extract contact info
    if website:
        website = clean_url(website)
        if not website or website == "-":
            return "", "", "", found_website

        print(f"    Fetching homepage...", end="", flush=True)
        try:
            status, html = fetch_page(website)
            if status == 200 and html:
                soup = BeautifulSoup(html, "lxml")
                emails, phones = extract_from_soup(soup, website)
                if emails:
                    email = pick_best_email(emails)
                if phones and not phone:
                    phone = phones[0]
                print(f" email={email or 'none'}, phone={phone or 'none'}")

                # If no email found, check contact pages
                if not email:
                    print(f"    Checking contact pages...", end="", flush=True)
                    contact_urls = find_contact_page_urls(get_base_url(website), soup)
                    for cu in contact_urls[:5]:
                        try:
                            cs, chtml = fetch_page(cu)
                            if cs == 200 and chtml:
                                csoup = BeautifulSoup(chtml, "lxml")
                                cemails, cphones = extract_from_soup(csoup, cu)
                                if cemails and not email:
                                    email = pick_best_email(cemails)
                                if cphones and not phone:
                                    phone = cphones[0]
                                if email:
                                    print(f" found email on {cu.split('/')[-1]}")
                                    break
                        except Exception:
                            continue
                    if not email:
                        print(f" none found")
            else:
                print(f" no HTML content")
        except Exception as e:
            print(f" error: {type(e).__name__}")

    return email, phone, website, found_website


def enrich(input_path, output_path):
    """Read scraped Excel, enrich missing Website/Phone/Email, write enriched Excel."""

    slug_base = os.path.splitext(os.path.basename(output_path))[0]
    checkpoint_file = f"/tmp/{slug_base}_enrich_checkpoint.json"
    temp_file = f"/tmp/{slug_base}_enrich_temp.xlsx"

    print("Loading Excel...")
    wb = load_workbook(input_path)
    ws = wb.active

    # Read all data
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c in range(1, ws.max_column + 1):
            row_data[headers[c - 1]] = str(ws.cell(r, c).value or "").strip()
        rows.append(row_data)

    print(f"Loaded {len(rows)} companies")

    # Load checkpoint
    done_slugs = load_checkpoint(checkpoint_file)

    # Process each company
    for idx, row in enumerate(rows):
        name = row.get("Company Name", "")
        existing_website = clean_url(row.get("Website", ""))
        existing_phone = row.get("Phone", "")

        # Slug: Türkçe karakterleri lower() ÖNCESİ dönüştür
        slug = name.strip()
        slug = slug.replace('İ', 'i').replace('I', 'i')
        slug = slug.replace('ı', 'i').replace('ş', 's').replace('Ş', 's')
        slug = slug.replace('ç', 'c').replace('Ç', 'c')
        slug = slug.replace('ğ', 'g').replace('Ğ', 'g')
        slug = slug.replace('ü', 'u').replace('Ü', 'u')
        slug = slug.replace('ö', 'o').replace('Ö', 'o')
        slug = slug.replace('â', 'a').replace('Â', 'a')
        slug = slug.lower()
        slug = re.sub(r'[^a-z0-9]', '-', slug)
        slug = re.sub(r'-+', '-', slug).strip('-')[:50]
        detail_url = row.get("Detail Page URL", "")
        url_hash = abs(hash(detail_url)) % 100000 if detail_url else idx
        slug = f"{slug}-{url_hash}"

        if slug in done_slugs:
            print(f"[{idx + 1}/{len(rows)}] SKIP (checkpoint): {name[:50]}")
            continue

        print(f"\n[{idx + 1}/{len(rows)}] {name[:60]}")

        email, phone, found_website_url, found_website = enrich_company(name, existing_website, existing_phone)

        # Update row data
        if email and is_valid_email(email):
            if not row.get("Email") and email != row.get("Email", ""):
                row["Email"] = email
                STATS["emails_added"] += 1
        if phone and phone != existing_phone:
            if not existing_phone:
                row["Phone"] = phone
                STATS["phones_added"] += 1
        if found_website and not existing_website and found_website_url:
            row["Website"] = found_website_url
            STATS["websites_added"] += 1

        STATS["companies_processed"] += 1
        done_slugs.add(slug)

        # Save checkpoint every 10 companies
        if (idx + 1) % 10 == 0:
            save_checkpoint(done_slugs, checkpoint_file)
            print(f"  *** Checkpoint at {idx + 1}/{len(rows)} ***")

        time.sleep(0.8)  # Rate limiting

    save_checkpoint(done_slugs, checkpoint_file)

    # Write enriched Excel
    print("\n\nWriting enriched Excel...")
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Exhibitors"

    out_headers = [
        "Company Name", "Website", "Phone", "Email", "Hall", "Stand", "Detail Page URL"
    ]
    ws_out.append(out_headers)

    for row in rows:
        ws_out.append([
            row.get("Company Name", ""),
            row.get("Website", ""),
            row.get("Phone", ""),
            row.get("Email", ""),
            row.get("Hall", ""),
            row.get("Stand", ""),
            row.get("Detail Page URL", ""),
        ])

    wb_out.save(temp_file)
    shutil.copy2(temp_file, output_path)
    print(f"Enriched Excel saved: {output_path}")

    # Calculate remaining
    remaining_no_email = sum(1 for r in rows if not r.get("Email"))
    remaining_no_phone = sum(1 for r in rows if not r.get("Phone"))

    print("\n" + "=" * 60)
    print(f"Companies processed:      {STATS['companies_processed']}")
    print(f"Emails added:             {STATS['emails_added']}")
    print(f"Phones added:             {STATS['phones_added']}")
    print(f"Websites added:           {STATS['websites_added']}")
    print(f"Remaining missing emails: {remaining_no_email}")
    print(f"Remaining missing phones: {remaining_no_phone}")
    print("=" * 60)


def main():
    if len(sys.argv) < 3:
        print("Usage: python enricher.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    enrich(sys.argv[1], sys.argv[2])


if __name__ == "__main__":
    main()
