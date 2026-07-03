import sys
import re
from urllib.parse import urlparse

from openpyxl import load_workbook


def valid_email(email):
    if not email:
        return False
    if not re.match(r'^[^@]+@[^@]+\.[^@]{2,}$', email):
        return False
    local, domain = email.split("@")
    if domain.lower() in ("example.com", "domain.com", "yourdomain.com", "email.com", "test.com"):
        return False
    if len(local) <= 1:
        return False
    return True


def valid_website_url(url):
    if not url:
        return False
    parsed = urlparse(url)
    return bool(parsed.scheme and parsed.netloc and "." in parsed.netloc)


def valid_phone(phone):
    if not phone:
        return False
    digits = re.sub(r'\D', '', phone)
    return 8 <= len(digits) <= 15


def valid_hall(hall):
    if not hall:
        return True
    return hall.isdigit()


def valid_stand(stand):
    if not stand:
        return True
    return bool(re.match(r'^[A-Za-z0-9]+$', stand))


def valid_detail_url(url):
    if not url:
        return False
    parsed = urlparse(url)
    return bool(parsed.scheme and parsed.netloc)


def fix_email(email):
    fixed = email.strip().lower().rstrip(".")
    return fixed if valid_email(fixed) else None


def fix_phone(phone):
    digits = re.sub(r"\D", "", phone)
    if len(digits) < 8 or len(digits) > 15:
        return None

    if len(digits) == 10:
        return f"+90 {digits[:3]} {digits[3:6]} {digits[6:8]} {digits[8:]}"
    elif len(digits) == 11 and digits.startswith("0"):
        d = digits[1:]
        return f"+90 {d[:3]} {d[3:6]} {d[6:8]} {d[8:]}"
    elif len(digits) == 12 and digits.startswith("90"):
        d = digits[2:]
        return f"+90 {d[:3]} {d[3:6]} {d[6:8]} {d[8:]}"
    elif len(digits) == 13 and digits.startswith("90"):
        d = digits[2:]
        return f"+90 {d[:3]} {d[3:6]} {d[6:8]} {d[8:]}"
    else:
        return f"+{digits}"


def normalize_domain(url):
    try:
        domain = urlparse(url).netloc.lower()
        return domain[4:] if domain.startswith("www.") else domain
    except Exception:
        return ""


def collect_format_issues(rows):
    issues = []

    for idx, row in enumerate(rows):
        name = row.get("Company Name", "")
        website = row.get("Website", "")
        phone = row.get("Phone", "")
        email = row.get("Email", "")
        hall = row.get("Hall", "")
        stand = row.get("Stand", "")
        detail_url = row.get("Detail Page URL", "")
        rn = idx + 2

        if email and not valid_email(email):
            issues.append(("email", rn, name, email))

        if website and not valid_website_url(website):
            issues.append(("website_format", rn, name, website))

        if phone and not valid_phone(phone):
            issues.append(("phone", rn, name, phone))

        if hall and not valid_hall(hall):
            issues.append(("hall", rn, name, hall))

        if stand and not valid_stand(stand):
            issues.append(("stand", rn, name, stand))

        if detail_url and not valid_detail_url(detail_url):
            issues.append(("detail_url", rn, name, detail_url))

    return issues


def collect_duplicates(rows):
    dupe_groups = []

    buckets = {
        "company": {},
        "website": {},
        "email": {},
        "detail_url": {},
    }

    for idx, row in enumerate(rows):
        name = row.get("Company Name", "")
        website = row.get("Website", "")
        email = row.get("Email", "")
        hall = row.get("Hall", "")
        stand = row.get("Stand", "")
        detail_url = row.get("Detail Page URL", "")
        rn = idx + 2

        if name:
            key = name.strip().lower()
            buckets["company"].setdefault(key, []).append((rn, name))

        if website:
            key = normalize_domain(website)
            if key:
                buckets["website"].setdefault(key, []).append((rn, website))

        if email:
            key = email.strip().lower()
            buckets["email"].setdefault(key, []).append((rn, email))

        if detail_url:
            key = detail_url.strip()
            buckets["detail_url"].setdefault(key, []).append((rn, detail_url))

    for label, bucket in buckets.items():
        for key, entries in bucket.items():
            if len(entries) > 1:
                dupe_groups.append((label, key, entries))

    return dupe_groups


FIXABLE = {"email": fix_email, "phone": fix_phone}


def apply_fixes(rows, issues):
    error_fixes = []
    normalizations = {"phone": 0, "email": 0}

    # Fix flagged issues
    for cat, rn, name, val in issues:
        fixer = FIXABLE.get(cat)
        if not fixer:
            continue
        fixed = fixer(val)
        if fixed is not None and fixed != val:
            row_idx = rn - 2
            if 0 <= row_idx < len(rows):
                col = "Email" if cat == "email" else "Phone"
                rows[row_idx][col] = fixed
                error_fixes.append((cat, rn, name, val, fixed))

    # Normalize all phone numbers (format consistently)
    for idx, row in enumerate(rows):
        phone = row.get("Phone", "")
        if phone:
            normed = fix_phone(phone)
            if normed and normed != phone:
                row["Phone"] = normed
                normalizations["phone"] += 1

    # Normalize all emails (lowercase, strip)
    for idx, row in enumerate(rows):
        email = row.get("Email", "")
        if email:
            normed = fix_email(email)
            if normed and normed != email:
                row["Email"] = normed
                normalizations["email"] += 1

    return error_fixes, normalizations


def print_report(filepath, rows, issues, dupe_groups, error_fixes=None, normalizations=None):
    total = len(rows)
    total_format = len(issues)
    total_dupe_rows = sum(len(e[2]) for e in dupe_groups)

    print("=" * 60)

    fmt_issues = {}
    for cat, rn, name, val in issues:
        fmt_issues.setdefault(cat, []).append((rn, name, val))

    fmt_labels = {
        "email": "Email",
        "website_format": "Website (URL)",
        "phone": "Phone",
        "hall": "Hall",
        "stand": "Stand",
        "detail_url": "Detail Page URL",
    }

    print(" FORMAT ISSUES")
    print("-" * 60)
    any_fmt = False
    for cat in ["email", "website_format", "phone", "hall", "stand", "detail_url"]:
        items = fmt_issues.get(cat, [])
        label = fmt_labels[cat]
        if items:
            any_fmt = True
            print(f"  {len(items):>4}  {label}")
            for rn, name, val in items[:10]:
                print(f"         Row {rn}: {val}")
            if len(items) > 10:
                print(f"         ... and {len(items) - 10} more")
    if not any_fmt:
        print("  (none)")

    print()
    print(" DUPLICATES")
    print("-" * 60)
    dupe_labels = {
        "company": "Company Name",
        "website": "Website domain",
        "email": "Email",
        "detail_url": "Detail Page URL",
    }
    any_dupe = False
    for label, key, entries in dupe_groups:
        any_dupe = True
        rows_str = ", ".join(f"Row {r[0]}" for r in entries)
        print(f"  {len(entries)}x  {dupe_labels[label]}: {key}")
        print(f"         {rows_str}")
    if not any_dupe:
        print("  (none)")

    if error_fixes:
        print()
        print(" FIXED ERRORS")
        print("-" * 60)
        for cat, rn, name, old, new in error_fixes:
            label = "Email" if cat == "email" else "Phone"
            print(f"  Row {rn}: {label}")
            print(f"         Old: {old}")
            print(f"         New: {new}")

    if normalizations and (normalizations["phone"] or normalizations["email"]):
        print()
        print(" NORMALIZED")
        print("-" * 60)
        parts = []
        if normalizations["phone"]:
            parts.append(f"Phone: {normalizations['phone']}")
        if normalizations["email"]:
            parts.append(f"Email: {normalizations['email']}")
        print("  " + ", ".join(parts))
        print()

    print(" SUMMARY")
    print("-" * 60)
    total_fixed = len(error_fixes) if error_fixes else 0
    remaining = total_format - total_fixed
    print(f"  Rows checked:       {total}")
    print(f"  Format issues:      {total_format}")
    print(f"  Errors fixed:       {total_fixed}")
    print(f"  Normalized:         {sum(normalizations.values()) if normalizations else 0}")
    print(f"  Duplicate groups:   {len(dupe_groups)} ({total_dupe_rows} rows)")
    print(f"  Remaining errors:   {max(remaining, 0)}")
    print("=" * 60)

    return total_format > 0 or len(dupe_groups) > 0


HEADERS_EXPECTED = [
    "Company Name", "Website", "Phone", "Email", "Hall", "Stand", "Detail Page URL",
]


def main():
    if len(sys.argv) < 2:
        print("Usage: python checker.py <enriched.xlsx> [--fix]")
        sys.exit(1)

    filepath = sys.argv[1]
    do_fix = "--fix" in sys.argv

    try:
        wb = load_workbook(filepath)
    except FileNotFoundError:
        print(f"Error: no such file — {filepath}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: cannot open Excel — {e}")
        sys.exit(1)

    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    missing = [h for h in HEADERS_EXPECTED if h not in headers]
    if missing:
        print(f"Error: missing columns — {', '.join(missing)}")
        wb.close()
        sys.exit(1)

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            row[headers[c - 1]] = str(val).strip() if val is not None else ""
        if any(row.values()):
            rows.append(row)

    wb.close()

    issues = collect_format_issues(rows)
    dupe_groups = collect_duplicates(rows)

    error_fixes = None
    normalizations = None
    if do_fix:
        error_fixes, normalizations = apply_fixes(rows, issues)

    has_problems = print_report(filepath, rows, issues, dupe_groups, error_fixes, normalizations)

    if do_fix:
        base, ext = filepath.rsplit(".", 1)
        out_path = f"{base}-fixed.{ext}"
        from openpyxl import Workbook
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Exhibitors"
        ws_out.append(HEADERS_EXPECTED)
        for row in rows:
            ws_out.append([row.get(h, "") for h in HEADERS_EXPECTED])
        wb_out.save(out_path)
        print(f"  Fixed file saved: {out_path}")

    sys.exit(1 if has_problems else 0)


if __name__ == "__main__":
    main()
